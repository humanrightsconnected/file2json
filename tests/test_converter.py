"""
Tests for the file2json converter module.
"""

import json
import os
import tempfile
from pathlib import Path

import pytest
import pandas as pd

from file2json.converter import detect_file_type, read_file, convert_to_json


# Fixture for creating test files
@pytest.fixture
def test_files():
    """Create temporary test files for different formats."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        # Create a CSV file
        csv_path = Path(tmp_dir) / "test.csv"
        df = pd.DataFrame({
            'name': ['Alice', 'Bob', 'Charlie'],
            'age': [25, 30, 35],
            'city': ['New York', 'London', 'Paris']
        })
        df.to_csv(csv_path, index=False)
        
        # Create a TSV file
        tsv_path = Path(tmp_dir) / "test.tsv"
        df.to_csv(tsv_path, sep='\t', index=False)
        
        # Create an Excel file
        excel_path = Path(tmp_dir) / "test.xlsx"
        with pd.ExcelWriter(excel_path) as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            df2 = pd.DataFrame({
                'product': ['Laptop', 'Phone', 'Tablet'],
                'price': [1000, 800, 500],
                'stock': [10, 20, 15]
            })
            df2.to_excel(writer, sheet_name='Sheet2', index=False)
        
        # Create a text file
        text_path = Path(tmp_dir) / "test.txt"
        with open(text_path, 'w') as f:
            f.write("Line 1\nLine 2\nLine 3\n")
        
        # Create a JSON file
        json_path = Path(tmp_dir) / "test.json"
        with open(json_path, 'w') as f:
            json.dump(df.to_dict(orient='records'), f)
        
        yield {
            'csv': csv_path,
            'tsv': tsv_path,
            'excel': excel_path,
            'text': text_path,
            'json': json_path,
            'temp_dir': tmp_dir
        }


def test_detect_file_type(test_files):
    """Test file type detection by extension."""
    assert detect_file_type(test_files['csv']) == 'csv'
    assert detect_file_type(test_files['tsv']) == 'tsv'
    assert detect_file_type(test_files['excel']) == 'excel'
    assert detect_file_type(test_files['text']) == 'text'
    assert detect_file_type(test_files['json']) == 'json'


def test_read_file_csv(test_files):
    """Test reading CSV files."""
    data = read_file(test_files['csv'])
    assert isinstance(data, list)
    assert len(data) == 3
    assert data[0]['name'] == 'Alice'
    assert data[0]['age'] == 25
    assert data[0]['city'] == 'New York'


def test_read_file_tsv(test_files):
    """Test reading TSV files."""
    data = read_file(test_files['tsv'])
    assert isinstance(data, list)
    assert len(data) == 3
    assert data[1]['name'] == 'Bob'
    assert data[1]['age'] == 30
    assert data[1]['city'] == 'London'


def test_read_file_excel(test_files):
    """Test reading Excel files with multiple sheets."""
    data = read_file(test_files['excel'])
    assert isinstance(data, dict)
    assert 'Sheet1' in data
    assert 'Sheet2' in data
    assert len(data['Sheet1']) == 3
    assert len(data['Sheet2']) == 3
    assert data['Sheet1'][2]['name'] == 'Charlie'
    assert data['Sheet2'][0]['product'] == 'Laptop'


def test_read_file_text(test_files):
    """Test reading text files."""
    data = read_file(test_files['text'])
    assert isinstance(data, dict)
    assert 'lines' in data
    assert len(data['lines']) == 3
    assert data['lines'][0] == 'Line 1'
    assert data['lines'][1] == 'Line 2'
    assert data['lines'][2] == 'Line 3'


def test_read_file_json(test_files):
    """Test reading JSON files."""
    data = read_file(test_files['json'])
    assert isinstance(data, list)
    assert len(data) == 3
    assert data[0]['name'] == 'Alice'
    assert data[1]['name'] == 'Bob'
    assert data[2]['name'] == 'Charlie'


def test_convert_to_json_string():
    """Test converting data to JSON string."""
    data = [{'name': 'Alice', 'age': 25}]
    json_str = convert_to_json(data)
    assert isinstance(json_str, str)
    assert json.loads(json_str) == data


def test_convert_to_json_file(test_files):
    """Test converting data to JSON file."""
    data = [{'name': 'Alice', 'age': 25}]
    output_path = os.path.join(test_files['temp_dir'], 'output.json')
    
    result = convert_to_json(data, output_path)
    assert "JSON saved to" in result
    
    # Verify file was created and contains correct data
    assert os.path.exists(output_path)
    with open(output_path, 'r') as f:
        saved_data = json.load(f)
    assert saved_data == data


def test_force_file_type(test_files):
    """Test forcing a specific file type."""
    # Force CSV parser on TSV file (will fail without force_type)
    with pytest.raises(Exception):
        read_file(test_files['tsv'], 'csv')
    
    # Force TSV parser on CSV file (should work)
    data = read_file(test_files['csv'], 'tsv')
    assert isinstance(data, list)
    assert len(data) == 3


def test_merged_cells_excel():
    """Test handling of merged cells in Excel files."""
    # Create a temporary Excel file with merged cells
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp:
        temp_path = temp.name
        
        # Create a workbook with merged cells
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add headers
        ws.cell(row=1, column=1).value = "Category"
        ws.cell(row=1, column=2).value = "Product"
        ws.cell(row=1, column=3).value = "Price"
        
        # Add data with merged cells in first column
        ws.cell(row=2, column=1).value = "Electronics"  # This value should apply to rows 2-4
        ws.cell(row=2, column=2).value = "Laptop"
        ws.cell(row=2, column=3).value = 1000
        
        ws.cell(row=3, column=2).value = "Phone"
        ws.cell(row=3, column=3).value = 800
        
        ws.cell(row=4, column=2).value = "Tablet"
        ws.cell(row=4, column=3).value = 500
        
        # Merge the cells in the Category column
        ws.merge_cells('A2:A4')
        
        # Add a different category
        ws.cell(row=5, column=1).value = "Books"  # This should apply to rows 5-6
        ws.cell(row=5, column=2).value = "Fiction"
        ws.cell(row=5, column=3).value = 20
        
        ws.cell(row=6, column=2).value = "Non-Fiction"
        ws.cell(row=6, column=3).value = 25
        
        # Merge the cells for the second category
        ws.merge_cells('A5:A6')
        
        # Save the workbook
        wb.save(temp_path)
    
    try:
        # Use our function to read the Excel file with merged cells
        data = read_file(temp_path, 'excel')
        
        # Verify the data
        sheet_data = data['Sheet']
        
        # Check that the merged cell values were propagated
        assert sheet_data[0]['Category'] == 'Electronics'
        assert sheet_data[1]['Category'] == 'Electronics'
        assert sheet_data[2]['Category'] == 'Electronics'
        assert sheet_data[3]['Category'] == 'Books'
        assert sheet_data[4]['Category'] == 'Books'
        
        # Check other values are correct
        assert sheet_data[0]['Product'] == 'Laptop'
        assert sheet_data[1]['Product'] == 'Phone'
        assert sheet_data[2]['Product'] == 'Tablet'
        assert sheet_data[3]['Product'] == 'Fiction'
        assert sheet_data[4]['Product'] == 'Non-Fiction'
        
    finally:
        # Clean up
        os.unlink(temp_path)