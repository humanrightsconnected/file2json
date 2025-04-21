"""
File to JSON converter module.
Handles detection, reading, and conversion of various file formats to JSON.
"""

import os
import json
import pandas as pd
from typing import Dict, List, Union, Optional, Any


def detect_file_type(file_path: str) -> str:
    """
    Detect file type based on extension or content.
    
    Args:
        file_path: Path to the input file
        
    Returns:
        String representing the detected file type
    """
    extension = os.path.splitext(file_path)[1].lower()
    
    # Check by extension first
    if extension in ['.xlsx', '.xls', '.xlsm', '.xlsb', '.odf', '.ods', '.odt']:
        return 'excel'
    elif extension == '.csv':
        return 'csv'
    elif extension == '.tsv':
        return 'tsv'
    elif extension == '.json':
        return 'json'
    elif extension in ['.txt', '.text']:
        return 'text'
    else:
        # Try to infer from content
        try:
            # Try reading as CSV first
            pd.read_csv(file_path, nrows=5)
            return 'csv'
        except Exception:
            try:
                # Try reading as TSV
                pd.read_csv(file_path, sep='\t', nrows=5)
                return 'tsv'
            except Exception:
                try:
                    # Try reading as Excel
                    pd.read_excel(file_path, nrows=5)
                    return 'excel'
                except Exception:
                    return 'unknown'


def read_file(file_path: str, file_type: Optional[str] = None) -> Any:
    """
    Read file based on detected or specified type.
    
    Args:
        file_path: Path to the input file
        file_type: Optional file type to force (excel, csv, tsv, json, text)
        
    Returns:
        Python object representation of the file content
        
    Raises:
        ValueError: If file type is unsupported or undetected
    """
    if not file_type:
        file_type = detect_file_type(file_path)
    
    if file_type == 'excel':
        # Read all sheets
        excel_file = pd.ExcelFile(file_path)
        data = {}
        
        # Use openpyxl to get merged cell information
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Process merged cells for this sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get all merged cell ranges
                merged_cells = ws.merged_cells.ranges
                
                # Handle merged cells that span rows
                for merged_range in merged_cells:
                    # Get merged cell value (from top-left cell)
                    top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    if top_left_cell.value is None:
                        continue
                    
                    # Only process if this is a column-wise merge (same column, multiple rows)
                    if merged_range.min_col == merged_range.max_col:
                        col_idx = merged_range.min_col - 1  # Convert to 0-based indexing
                        
                        # Skip if the column index is out of bounds for our dataframe
                        if col_idx >= len(df.columns):
                            continue
                        
                        col_name = df.columns[col_idx]
                        
                        # For each row in the merged range, apply the value
                        for row_idx in range(merged_range.min_row - 1, merged_range.max_row):
                            # Skip if row index is out of bounds
                            if row_idx >= len(df):
                                continue
                            
                            # Apply the merged cell value to this row
                            df.at[row_idx, col_name] = top_left_cell.value
            
            data[sheet_name] = df.to_dict(orient='records')
        return data
    
    elif file_type == 'csv':
        df = pd.read_csv(file_path)
        return df.to_dict(orient='records')
    
    elif file_type == 'tsv':
        df = pd.read_csv(file_path, sep='\t')
        return df.to_dict(orient='records')
    
    elif file_type == 'json':
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    elif file_type == 'text':
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        return {"lines": [line.rstrip('\n') for line in lines]}
    
    else:
        raise ValueError(f"Unsupported or undetected file type for {file_path}")


def convert_to_json(data: Any, output_path: Optional[str] = None) -> str:
    """
    Convert data to JSON and optionally save to file.
    
    Args:
        data: Python object to convert to JSON
        output_path: Optional path to save JSON output
        
    Returns:
        JSON string or confirmation message if saved to file
    """
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return f"JSON saved to {output_path}"
    else:
        return json.dumps(data, indent=2, ensure_ascii=False)